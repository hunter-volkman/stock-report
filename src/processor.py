import os
import subprocess
import shutil
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
from dateutil import tz
from viam.logging import getLogger

LOGGER = getLogger(__name__)

class WorkbookProcessor:
    """
    WorkbookProcessor class that handles the creation of daily fill reports
    by processing raw data, updating Excel templates, and fixing XML-level formulas and row counts.
    """
    def __init__(self, work_dir, export_script, api_key_id, api_key, org_id, 
                 timezone="America/New_York", export_start_time_weekday="7:00", 
                 export_end_time_weekday="19:00", export_start_time_weekend="8:00", 
                 export_end_time_weekend="16:00"):
        """
        Initialize the WorkbookProcessor with configuration parameters.

        Args:
            work_dir (str): Directory where workbooks and temporary files are stored.
            export_script (str): Path to the vde.py export script.
            api_key_id (str): Viam API key ID for authentication.
            api_key (str): Viam API key for authentication.
            org_id (str): Viam organization ID.
            timezone (str): Timezone for date handling (default: "America/New_York").
            export_start_time_weekday (str): Start time for weekday exports (e.g., "7:00").
            export_end_time_weekday (str): End time for weekday exports (e.g., "19:00").
            export_start_time_weekend (str): Start time for weekend exports (e.g., "8:00").
            export_end_time_weekend (str): End time for weekend exports (e.g., "16:00").
        """
        self.work_dir = work_dir
        self.export_script_path = export_script
        self.export_script_dir = os.path.dirname(export_script)
        self.api_key_id = api_key_id
        self.api_key = api_key
        self.org_id = org_id
        self.timezone = timezone
        self.export_start_time_weekday = export_start_time_weekday
        self.export_end_time_weekday = export_end_time_weekday
        self.export_start_time_weekend = export_start_time_weekend
        self.export_end_time_weekend = export_end_time_weekend
        
        # Check for viam-python-data-export virtual environment
        self.venv_path = os.path.join(self.export_script_dir, ".venv")
        if not os.path.exists(self.venv_path):
            LOGGER.warning(f"viam-python-data-export virtual environment not found at {self.venv_path}, attempting to set it up")
            self._setup_venv()
        else:
            LOGGER.info(f"Found viam-python-data-export virtual environment at {self.venv_path}")
    
    def _setup_venv(self):
        """Set up the virtual environment for viam-python-data-export if it doesn't exist"""
        try:
            setup_script = os.path.join(self.export_script_dir, "setup.sh")
            if os.path.exists(setup_script):
                LOGGER.info(f"Running setup script for viam-python-data-export: {setup_script}")
                subprocess.run(
                    ["bash", "-c", f"cd {self.export_script_dir} && source ./setup.sh"],
                    check=True,
                    shell=False
                )
                LOGGER.info("viam-python-data-export setup script completed successfully!")
            else:
                LOGGER.error(f"Setup script not found at {setup_script}")
        except Exception as e:
            LOGGER.error(f"Failed to set up viam-python-data-export virtual environment: {e}")

    def get_yesterday_date(self):
        """Get yesterday's date in the configured timezone."""
        now = datetime.now(tz.gettz(self.timezone))
        return now - timedelta(days=1)

    def _get_export_times_for_day(self, target_date):
        """Determine export start and end times based on whether it's a weekday or weekend."""
        is_weekday = target_date.weekday() < 5  # Mon=0, Sun=6
        start_time_str = self.export_start_time_weekday if is_weekday else self.export_start_time_weekend
        end_time_str = self.export_end_time_weekday if is_weekday else self.export_end_time_weekend
        return start_time_str, end_time_str, is_weekday

    def export_raw_data(self, output_file, target_date=None):
        """
        Run the vde.py script to export raw data for the specified date.
        
        Args:
            output_file (str): Path where the raw export file should be saved.
            target_date (datetime, optional): Date to export data for (defaults to yesterday).
            
        Returns:
            tuple: (output_file_path, is_weekday) where output_file_path is the path to the exported file
                   and is_weekday is a boolean indicating if the target_date is a weekday.
        """
        # Use the provided date or default to yesterday
        if target_date is None:
            target_date = self.get_yesterday_date()
            LOGGER.info(f"No target date provided, using yesterday: {target_date.strftime('%Y-%m-%d')}")
        
        # Get the appropriate export times based on the day
        start_time_str, end_time_str, is_weekday = self._get_export_times_for_day(target_date)
        
        # Parse the time strings into hours and minutes
        start_hour, start_minute = map(int, start_time_str.split(':'))
        end_hour, end_minute = map(int, end_time_str.split(':'))
        
        # Create the datetime objects for start and end times
        start_time = target_date.replace(hour=start_hour, minute=start_minute, second=0, microsecond=0)
        end_time = target_date.replace(hour=end_hour, minute=end_minute, second=0, microsecond=0)

        LOGGER.info(f"Exporting data from {start_time} to {end_time} ({start_time_str} to {end_time_str})")

        # Construct the shell script to run vde.py with its virtual environment
        venv_python = os.path.join(self.venv_path, "bin", "python")
        
        # Build the command        
        cmd = [
            venv_python,
            self.export_script_path,
            "-vv", 
            "excel",
            "--apiKeyId", self.api_key_id,
            "--apiKey", self.api_key,
            "--orgId", self.org_id,
            "--resourceName", "langer_fill",
            "--start", start_time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "--end", end_time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "--timezone", self.timezone,
            "--bucketPeriod", "PT5M",
            "--bucketMethod", "pct99",
            "--includeKeys", ".*_raw",
            "--output", output_file,
            "--tab", "RAW"
        ]

        # Create a masked version of the command for logging
        cmd_mask = cmd.copy()

        # Find the index of the sensitive parameters and mask them
        if "--apiKeyId" in cmd_mask:
            idx = cmd_mask.index("--apiKeyId")
            if idx + 1 < len(cmd_mask):
                cmd_mask[idx + 1] = "<redacted>"
        if "--apiKey" in cmd_mask:
            idx = cmd_mask.index("--apiKey")
            if idx + 1 < len(cmd_mask):
                cmd_mask[idx + 1] = "<redacted>"

        try:
            # Log the masked command
            LOGGER.info(f"Running vde.py command: {' '.join(cmd_mask)}")
            process = subprocess.run(
                cmd, 
                check=True, 
                cwd=self.export_script_dir,
                capture_output=True,
                text=True
            )
            
            # Log a summary of the output
            stdout_lines = process.stdout.strip().split('\n')
            if stdout_lines:
                # Just log a few lines to avoid excessive logging
                if len(stdout_lines) > 4:
                    LOGGER.info("vde.py output (sample):")
                    for line in stdout_lines[:2]:
                        LOGGER.info(f"  {line}")
                    LOGGER.info("  ...")
                    for line in stdout_lines[-2:]:
                        LOGGER.info(f"  {line}")
                else:
                    LOGGER.info("vde.py output:")
                    for line in stdout_lines:
                        LOGGER.info(f"  {line}")
            
            if not os.path.exists(output_file):
                raise FileNotFoundError("vde.py ran but raw_export.xlsx was not created.")
            
            LOGGER.info(f"Generated raw data at {output_file}")
            return output_file, is_weekday
        except subprocess.CalledProcessError as e:
            LOGGER.error(f"Failed to run vde.py: {e}")
            if e.stderr:
                stderr_lines = e.stderr.strip().split('\n')
                LOGGER.error("vde.py stderr output:")
                for line in stderr_lines:
                    LOGGER.error(f"  {line}")
            raise RuntimeError(f"vde.py export failed: {e}")

    def _update_raw_import_sheet(self, raw_file, output_file):
        """
        Updates the Raw Import sheet in the output file with data from the raw file.
        
        Args:
            raw_file (str): Path to the raw export Excel file.
            output_file (str): Path to the output workbook.
            
        Returns:
            int: Number of data rows copied to the Raw Import sheet.
        """
        try:
            # Load data from raw export file
            LOGGER.info(f"Loading raw data from {raw_file}")
            raw_wb = openpyxl.load_workbook(raw_file, data_only=True)
            
            if "RAW" not in raw_wb.sheetnames:
                LOGGER.error("RAW sheet not found in exported data")
                raise ValueError("RAW sheet not found in exported data")
                
            raw_sheet = raw_wb["RAW"]
            
            # Get headers and data from raw sheet
            headers = [cell.value for cell in raw_sheet[1]]
            data_rows = list(raw_sheet.iter_rows(min_row=2, values_only=True))
            
            LOGGER.info(f"Loaded {len(data_rows)} rows of data from raw export")
            
            # Open the output workbook
            LOGGER.info(f"Opening output workbook: {output_file}")
            output_wb = openpyxl.load_workbook(output_file)
            
            if "Raw Import" not in output_wb.sheetnames:
                LOGGER.error("Raw Import sheet not found in template")
                raise ValueError("Raw Import sheet not found in template")
                
            output_sheet = output_wb["Raw Import"]
            
            # Clear existing data from Raw Import sheet (keeping headers)
            LOGGER.info("Clearing existing data from Raw Import sheet")
            for row in output_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            
            # Copy data to Raw Import sheet
            LOGGER.info("Copying data to Raw Import sheet")
            for r_idx, row_data in enumerate(data_rows, start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    output_sheet.cell(row=r_idx, column=c_idx).value = value
            
            # Save the workbook
            LOGGER.info(f"Saving updated workbook to {output_file}")
            output_wb.save(output_file)
            
            LOGGER.info(f"Raw Import sheet updated with {len(data_rows)} rows of data")
            return len(data_rows)  # Return the number of data rows for use in fixing
            
        except Exception as e:
            LOGGER.error(f"Error updating Raw Import sheet: {e}")
            raise

    def process(self, target_date=None):
        """
        Main processing function: export the data and update the template to create a WIP file.
        
        Args:
            target_date (datetime, optional): Date to process (defaults to yesterday).
            
        Returns:
            tuple: (wip_path, num_data_rows) where wip_path is the path to the intermediate workbook
                   and num_data_rows is the number of data rows in the Raw Import sheet.
        """
        os.makedirs(self.work_dir, exist_ok=True)
        
        if target_date is None:
            target_date = self.get_yesterday_date()
            
        LOGGER.info(f"Starting workbook processing for date: {target_date.strftime('%Y-%m-%d')}")
        
        try:
            # 1. Export raw data
            raw_file = os.path.join(self.work_dir, "raw_export.xlsx")
            raw_file, is_weekday = self.export_raw_data(raw_file, target_date)
            
            # 2. Determine template and output filenames
            # Using a single template file regardless of weekday/weekend
            template_path = os.path.join(self.work_dir, "template.xlsx")
            
            output_filename = f"3895th_{target_date.strftime('%m%d%y')}_wip.xlsx"  # Intermediate WIP file
            output_path = os.path.join(self.work_dir, output_filename)
            
            # Check if template exists
            if not os.path.exists(template_path):
                LOGGER.error(f"Template file not found: {template_path}")
                raise FileNotFoundError(f"Template file not found: {template_path}")
            
            # 3. Copy the template to create the new workbook
            LOGGER.info(f"Creating new workbook from template: {template_path}")
            shutil.copy(template_path, output_path)
            
            # 4. Update the Raw Import sheet and get the number of data rows
            num_data_rows = self._update_raw_import_sheet(raw_file, output_path)
            
            # 5. Return the WIP path and number of data rows
            LOGGER.info(f"Successfully created WIP workbook: {output_path}")
            return output_path, num_data_rows
            
        except Exception as e:
            LOGGER.error(f"Error processing workbook: {e}")
            raise

    def get_sheet_mappings(self, excel_path):
        """
        Extracts the correct mapping of sheet names to their actual XML filenames.
        
        Args:
            excel_path (str): Path to the Excel file.

        Returns:
            dict: Mapping of sheet names to actual worksheet XML file names (e.g., "Sorted Raw" -> "sheet5.xml").
        """
        temp_dir = os.path.join(self.work_dir, "temp_excel")
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            with zipfile.ZipFile(excel_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)

            workbook_xml_path = os.path.join(temp_dir, "xl", "workbook.xml")
            rels_xml_path = os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels")

            # Check if the necessary XML files exist
            if not os.path.exists(workbook_xml_path):
                raise FileNotFoundError(f"workbook.xml not found in {excel_path}")
            if not os.path.exists(rels_xml_path):
                raise FileNotFoundError(f"workbook.xml.rels not found in {excel_path}")

            sheet_mapping = {}

            # Parse workbook.xml to get sheet names and their relationship IDs
            wb_tree = ET.parse(workbook_xml_path)
            wb_root = wb_tree.getroot()
            
            # Handle Excel namespaces properly
            ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheet_rel_map = {}  # Map r:id to sheet name
            for sheet in wb_root.findall(".//ns:sheets/ns:sheet", ns):
                sheet_name = sheet.attrib["name"]
                sheet_rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
                sheet_rel_map[sheet_rel_id] = sheet_name

            # Parse workbook.xml.rels to get sheet file paths
            rels_tree = ET.parse(rels_xml_path)
            rels_root = rels_tree.getroot()
            rels_ns = {'ns': 'http://schemas.openxmlformats.org/package/2006/relationships'}

            for rel in rels_root.findall(".//ns:Relationship", rels_ns):
                rel_id = rel.attrib["Id"]
                target = rel.attrib["Target"]
                
                if rel_id in sheet_rel_map and "worksheets" in target:
                    sheet_name = sheet_rel_map[rel_id]
                    sheet_mapping[sheet_name] = os.path.basename(target)

            return sheet_mapping
            
        except Exception as e:
            LOGGER.error(f"Error extracting sheet mappings: {e}")
            raise
        finally:
            # Cleanup - but don't remove if we're using this directory for fix operation
            if os.path.exists(temp_dir) and "temp_excel" not in excel_path:
                shutil.rmtree(temp_dir)

    def fix(self, wip_path, num_data_rows):
        """
        Improved fix method that updates Excel workbook XML properly.
        """
        # Define temp directory with a unique name to avoid conflicts
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        temp_dir = os.path.join(self.work_dir, f"temp_excel_{timestamp}")
        
        try:
            # Ensure WIP file exists
            if not os.path.exists(wip_path):
                LOGGER.error(f"WIP file not found: {wip_path}")
                raise FileNotFoundError(f"WIP file not found: {wip_path}")

            # Create a fresh temp directory
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            os.makedirs(temp_dir, exist_ok=True)
            LOGGER.info(f"Created temp directory: {temp_dir}")
            
            # Extract the Excel file
            LOGGER.info(f"Extracting WIP Excel file: {wip_path}")
            with zipfile.ZipFile(wip_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)
            
            # Get sheet mappings
            LOGGER.info("Obtaining sheet mappings")
            sheet_mappings = self.get_sheet_mappings(wip_path)
            LOGGER.info(f"Sheet mappings: {sheet_mappings}")
            
            # Define namespaces properly
            namespaces = {
                'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            
            # Register namespaces for proper XML generation
            for prefix, uri in namespaces.items():
                ET.register_namespace(prefix, uri)
            # Also register the default namespace
            ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
            
            # Path to worksheets directory
            worksheets_dir = os.path.join(temp_dir, "xl", "worksheets")
            if not os.path.exists(worksheets_dir):
                LOGGER.error(f"Worksheets directory not found: {worksheets_dir}")
                raise FileNotFoundError(f"Worksheets directory not found: {worksheets_dir}")
            
            # Process each sheet that needs fixing
            for sheet_name in ["Sorted Raw", "Calibrated Values", "Bounded Calibrated"]:
                if sheet_name not in sheet_mappings:
                    LOGGER.warning(f"Sheet '{sheet_name}' not found in workbook. Skipping...")
                    continue
                
                sheet_xml_path = os.path.join(worksheets_dir, sheet_mappings[sheet_name])
                if not os.path.exists(sheet_xml_path):
                    LOGGER.error(f"Sheet XML file not found: {sheet_xml_path}")
                    continue
                
                LOGGER.info(f"Processing sheet: {sheet_name} ({sheet_xml_path})")
                
                # Parse sheet XML
                try:
                    tree = ET.parse(sheet_xml_path)
                    root = tree.getroot()
                    
                    # Find sheetData element
                    sheet_data = root.find(".//ns:sheetData", namespaces)
                    if sheet_data is None:
                        LOGGER.warning(f"No sheetData found in {sheet_name}, skipping modifications")
                        continue
                    
                    # Remove excess rows
                    rows_to_remove = []
                    for row in sheet_data.findall(".//ns:row", namespaces):
                        row_number = int(row.attrib.get("r", "0"))
                        if row_number > num_data_rows + 1:  # +1 for header row
                            rows_to_remove.append(row)
                    
                    for row in rows_to_remove:
                        sheet_data.remove(row)
                        LOGGER.info(f"Removed excess row {row.attrib.get('r')} from {sheet_name}")
                    
                    # Handle the SORT formula for Sorted Raw sheet
                    if sheet_name == "Sorted Raw":
                        LOGGER.info(f"Updating SORT formula in {sheet_name}")
                        
                        # Find cell A2 where the SORT formula should be
                        cell_a2 = None
                        for row in sheet_data.findall(".//ns:row", namespaces):
                            if row.attrib.get("r") == "2":
                                for cell in row.findall(".//ns:c", namespaces):
                                    if cell.attrib.get("r") == "A2":
                                        cell_a2 = cell
                                        break
                        
                        # If A2 cell not found, we need to create it
                        if cell_a2 is None:
                            LOGGER.info("Cell A2 not found, creating it")
                            # Look for row 2
                            row_2 = None
                            for row in sheet_data.findall(".//ns:row", namespaces):
                                if row.attrib.get("r") == "2":
                                    row_2 = row
                                    break
                            
                            # If row 2 doesn't exist, create it
                            if row_2 is None:
                                LOGGER.info("Row 2 not found, creating it")
                                row_2 = ET.SubElement(sheet_data, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row")
                                row_2.set("r", "2")
                            
                            # Create cell A2
                            cell_a2 = ET.SubElement(row_2, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c")
                            cell_a2.set("r", "A2")
                        
                        # Update or create formula element
                        formula = cell_a2.find(".//ns:f", namespaces)
                        if formula is None:
                            formula = ET.SubElement(cell_a2, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}f")
                        
                        # Set the updated formula with the correct row count
                        sort_formula = f"_xlfn._xlws.SORT('Raw Import'!A2:X{num_data_rows + 1},1,1)"
                        formula.text = sort_formula
                        formula.set("t", "array")
                        formula.set("ref", f"A2:X{num_data_rows + 1}")
                        
                        # Ensure there's a value element
                        value = cell_a2.find(".//ns:v", namespaces)
                        if value is None:
                            value = ET.SubElement(cell_a2, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
                        value.text = "0"
                        
                        LOGGER.info(f"Updated SORT formula: {sort_formula}")
                    
                    # Save the modified sheet XML
                    tree.write(sheet_xml_path, encoding="UTF-8", xml_declaration=True)
                    LOGGER.info(f"Saved modifications to {sheet_xml_path}")
                    
                except Exception as e:
                    LOGGER.error(f"Error processing sheet {sheet_name}: {e}")
                    raise
            
            # Create the final file path
            final_path = wip_path.replace("_wip.xlsx", "_final.xlsx")
            LOGGER.info(f"Creating final Excel file: {final_path}")
            
            # Create the new zip file (Excel file)
            with zipfile.ZipFile(final_path, "w", zipfile.ZIP_DEFLATED) as zip_out:
                for root_dir, _, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zip_out.write(file_path, arcname)
            
            LOGGER.info(f"Successfully created final Excel file: {final_path}")
            return final_path
            
        except Exception as e:
            LOGGER.error(f"Error in fix method: {e}")
            raise
        finally:
            # Clean up temporary directory
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                LOGGER.info(f"Cleaned up temporary directory: {temp_dir}")