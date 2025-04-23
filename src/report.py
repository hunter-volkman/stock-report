import asyncio
import datetime
import json
import os
import base64
import fasteners
import shutil
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from typing import Mapping, Optional, Any, Dict, List
from viam.module.module import Module
from viam.components.sensor import Sensor
from viam.components.camera import Camera
from viam.proto.app.robot import ComponentConfig
from viam.resource.base import ResourceBase
from viam.resource.types import Model, ModelFamily
from viam.utils import SensorReading, struct_to_dict
from viam.logging import getLogger
from viam.media.video import ViamImage
from dateutil import tz
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import (
    Mail, Attachment, FileContent, FileName,
    FileType, Disposition, Email, Content
)
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO

from .export import DataExporter

LOGGER = getLogger(__name__)

class StockReportEmail(Sensor):
    MODEL = Model(ModelFamily("hunter", "stock-report"), "email")
    
    @classmethod
    def new(cls, config: ComponentConfig, dependencies: Mapping[str, ResourceBase]) -> "StockReportEmail":
        """Create a new StockReportEmail instance."""
        instance = cls(config.name)
        instance.reconfigure(config, dependencies)
        return instance
    
    @classmethod
    def validate_config(cls, config: ComponentConfig) -> list[str]:
        """Validate the configuration and return required dependencies."""
        if not config.attributes.fields["location"].string_value:
            raise ValueError("location must be specified")

        attributes = struct_to_dict(config.attributes)

        # Validate recipients
        recipients = attributes.get("recipients", [])
        if not recipients or not isinstance(recipients, list):
            raise ValueError("recipients must be a non-empty list of email addresses")

        # Validate send_time
        send_time = attributes.get("send_time", "20:00")
        try:
            datetime.datetime.strptime(str(send_time), "%H:%M")
        except ValueError:
            raise ValueError(f"Invalid send_time '{send_time}': must be in 'HH:MM' format")

        # Validate process_time if provided
        process_time = attributes.get("process_time")
        if process_time:
            try:
                datetime.datetime.strptime(str(process_time), "%H:%M")
            except ValueError:
                raise ValueError(f"Invalid process_time '{process_time}': must be in 'HH:MM' format")

        # Validate capture_times if provided
        if "capture_times" in attributes:
            for time_str in attributes["capture_times"]:
                try:
                    datetime.datetime.strptime(time_str, "%H:%M")
                except ValueError:
                    raise ValueError(f"Invalid capture_times entry '{time_str}': must be in 'HH:MM' format")
        
        # Validate store hours
        for hours_key in ["hours_weekdays", "hours_weekends"]:
            if hours_key not in attributes:
                raise ValueError(f"{hours_key} is required")
            hours = attributes[hours_key]
            if not isinstance(hours, list) or len(hours) != 2:
                raise ValueError(f"'{hours_key}' must be a list with two elements: [opening_time, closing_time]")

            # Validate each time string
            for time_str in hours:
                try:
                    datetime.datetime.strptime(str(time_str), "%H:%M")
                except ValueError:
                    raise ValueError(f"Invalid time format in '{hours_key}': '{time_str}' - must be in 'HH:MM' format")

        # Check SendGrid API key
        sendgrid_api_key = attributes.get("sendgrid_api_key", "")
        if not sendgrid_api_key:
            LOGGER.warning("No SendGrid API key provided in configuration")

       
        # Check API key ID
        api_key_id = attributes.get("api_key_id", "")
        if not api_key_id:
            LOGGER.warning("No API key ID provided in configuration")

        # Check API key
        api_key = attributes.get("api_key", "")
        if not api_key:
            LOGGER.warning("No API key provided in configuration")

        # Check org ID
        org_id = attributes.get("org_id", "")
        if not api_key:
            LOGGER.warning("No org ID provided in configuration")

        # Check camera configuration if enabled
        include_images = attributes.get("include_images", False)
        if include_images and not attributes.get("camera_name"):
            raise ValueError("camera_name must be specified when include_images is true")

        # Return required dependencies
        # TODO: Standardize code...
        deps = []

        # Add camera dependency if configured
        if include_images and attributes.get("camera_name"):
            camera_name = attributes.get("camera_name")
            # If camera_name includes a remote name, use it directly
            if ":" in camera_name:
                deps.append(camera_name)
            else:
                # No remote prefix needed now
                deps.append(camera_name)

        LOGGER.info(f"StockReportEmail.validate_config completed for {deps}")
        return deps
    
    def __init__(self, name: str):
        super().__init__(name)
        self.dependencies = {}
        self.config = None
        self.location = ""
        self.teleop_url = ""
        self.recipients = []

        # Email configuration
        self.sendgrid_api_key = ""
        self.sender_email = "no-reply@viam.com"
        self.sender_name = "Stock Report Module"

        # Camera configuration
        self.camera_name = ""
        self.include_images = False
        self.image_width = 640
        self.image_height = 480
        # Remove single capture times
        # self.capture_times = ["08:00", "10:00", "12:00", "14:00", "16:00", "18:00"]
        self.capture_times_weekday = ["07:00", "08:00", "10:00", "12:00", "14:00", "16:00", "18:00"]
        self.capture_times_weekend = ["08:00", "09:00", "11:00", "16:00"]
        self.last_capture_time = None

        # API configuration
        # Viam Data Client
        self.api_key_id = ""
        self.api_key = ""
        self.org_id = ""

        # Store hours defaults
        self.hours_weekdays = ["07:00", "19:30"]  # Default for weekdays (Mon-Fri)
        self.hours_weekends = ["08:00", "17:00"]  # Default for weekends (Sat-Sun)
        self.timezone = "America/New_York"

        # Simplified scheduling
        self.process_time = "20:00"
        self.send_time = "20:30"
        
        # State
        self.last_processed_time = None
        self.last_sent_time = None
        self.last_capture_time = None
        self.last_workbook_path = None
        self.total_reports_sent = 0
        self.report_status = "not_sent"
        self.workbook_status = "not_processed"

        # State persistence
        self.state_dir = os.path.join(os.path.expanduser("~"), ".stock-report")
        self.state_file = os.path.join(self.state_dir, f"{name}.json")
        self.workbooks_dir = os.path.join(self.state_dir, "workbooks")
        self.images_dir = os.path.join(self.state_dir, "images")
        os.makedirs(self.state_dir, exist_ok=True)
        os.makedirs(self.workbooks_dir, exist_ok=True)
        os.makedirs(self.images_dir, exist_ok=True)

        # Background tasks
        self._process_task = None
        self._send_task = None
        self._capture_task = None

        # Load state silently
        self._load_state()
    
    def _load_state(self):
        """Load persistent state from file with locking."""
        if os.path.exists(self.state_file):
            lock = fasteners.InterProcessLock(f"{self.state_file}.lock")
            try:
                if lock.acquire(blocking=True, timeout=5):
                    try:
                        with open(self.state_file, "r") as f:
                            state = json.load(f)
                            self.last_processed_time = (
                                datetime.datetime.fromisoformat(state["last_processed_time"])
                                if state.get("last_processed_time") else None
                            )
                            self.last_sent_time = (
                                datetime.datetime.fromisoformat(state["last_sent_time"])
                                if state.get("last_sent_time") else None
                            )
                            self.last_capture_time = (
                                datetime.datetime.fromisoformat(state["last_capture_time"])
                                if state.get("last_capture_time") else None
                            )
                            self.last_workbook_path = state.get("last_workbook_path")
                            self.total_reports_sent = state.get("total_reports_sent", 0)
                            self.report_status = state.get("report_status", "not_sent")
                            self.workbook_status = state.get("workbook_status", "not_processed")
                        LOGGER.info(f"Loaded state from {self.state_file}")
                    finally:
                        lock.release()
                else:
                    LOGGER.warning(f"Could not acquire lock to load state for {self.name}")
            except Exception as e:
                LOGGER.error(f"Error loading state: {e}")
        else:
            LOGGER.info(f"No state file at {self.state_file}, starting fresh")
    
    def _save_state(self):
        """Save state to file for persistence across restarts using file locking."""
        lock = fasteners.InterProcessLock(f"{self.state_file}.lock")
        try:
            if lock.acquire(blocking=True, timeout=5):
                try:
                    state = {
                        "last_processed_time": self.last_processed_time.isoformat() if self.last_processed_time else None,
                        "last_sent_time": self.last_sent_time.isoformat() if self.last_sent_time else None,
                        "last_capture_time": self.last_capture_time.isoformat() if self.last_capture_time else None,
                        "last_workbook_path": self.last_workbook_path,
                        "total_reports_sent": self.total_reports_sent,
                        "report_status": self.report_status,
                        "workbook_status": self.workbook_status
                    }
                    temp_file = f"{self.state_file}.tmp"
                    with open(temp_file, "w") as f:
                        json.dump(state, f)
                    os.replace(temp_file, self.state_file)
                    LOGGER.debug(f"Saved state to {self.state_file}")
                finally:
                    lock.release()
            else:
                LOGGER.warning(f"Could not acquire lock to save state for {self.name}")
        except Exception as e:
            LOGGER.error(f"Error saving state: {e}")
    
    def reconfigure(self, config: ComponentConfig, dependencies: Mapping[str, ResourceBase]):
        """Configure the stock report with updated settings."""
        # Store config for later use
        self.config = config
    
        # Configure from attributes
        self.location = config.attributes.fields["location"].string_value
        attributes = struct_to_dict(config.attributes)

        # Email configuration
        self.recipients = attributes.get("recipients", [])
        self.sender_email = attributes.get("sender_email", "no-reply@viam.com")
        self.sender_name = attributes.get("sender_name", "Stock Report Module")
        self.sendgrid_api_key = attributes.get("sendgrid_api_key", "")
        self.teleop_url = attributes.get("teleop_url", "")

        # API configuration
        self.api_key_id = attributes.get("api_key_id", "")
        self.api_key = attributes.get("api_key", "")
        self.org_id = attributes.get("org_id", "")

        # Image configuration
        self.include_images = attributes.get("include_images", False)
        if isinstance(self.include_images, str):
            self.include_images = self.include_images.lower() == "true"
        self.camera_name = attributes.get("camera_name", "")
        self.image_width = int(attributes.get("image_width", 640))
        self.image_height = int(attributes.get("image_height", 480))

        # Store hours
        self.hours_weekdays = attributes.get("hours_weekdays", ["07:00", "19:30"])
        self.hours_weekends = attributes.get("hours_weekends", ["08:00", "17:00"])

        # Scheduling configuration
        self.send_time = attributes.get("send_time", "20:00")
        # Default to send_time if not set
        self.process_time = attributes.get("process_time", self.send_time)  
        # If still unset, calculate 1 hour before send_time
        if not self.process_time:  
            send_dt = datetime.datetime.strptime(self.send_time, "%H:%M")
            process_dt = send_dt - datetime.timedelta(hours=1)
            self.process_time = process_dt.strftime("%H:%M")
        self.timezone = attributes.get("timezone", "America/New_York")
        # self.capture_times = attributes.get("capture_times", ["08:00", "10:00", "12:00", "14:00", "16:00", "18:00"])
        # New schedule configuration
        self.capture_times_weekday = attributes.get("capture_times_weekday", ["07:00", "08:00", "10:00", "12:00", "14:00", "16:00", "18:00"])
        self.capture_times_weekend = attributes.get("capture_times_weekend", ["08:00", "09:00", "11:00", "16:00"])
        
        # Sort the capture times to ensure they're in chronological order
        # self.capture_times = sorted(list(set(self.capture_times)))
        self.capture_times_weekday = sorted(list(set(self.capture_times_weekday)))
        self.capture_times_weekend = sorted(list(set(self.capture_times_weekend)))

        # Store dependencies
        self.dependencies = dependencies

        # Cancel existing tasks if they exist
        if self._process_task and not self._process_task.done():
            self._process_task.cancel()
        if self._send_task and not self._send_task.done():
            self._send_task.cancel()
        if self._capture_task and not self._capture_task.done():
            self._capture_task.cancel()

        # Log configuration details
        LOGGER.info(f"Configured {self.name} for location '{self.location}'")
        LOGGER.info(f"Process time: {self.process_time}, Send time: {self.send_time}")
        LOGGER.info(f"Will send reports to: {', '.join(self.recipients)}")
        if self.sendgrid_api_key:
            LOGGER.info("SendGrid API key configured")
        else:
            LOGGER.warning("No SendGrid API key configured")
        # Update logging
        if self.include_images:
            LOGGER.info(f"Will capture images from camera: {self.camera_name}")
            # LOGGER.info(f"Capture times: {', '.join(self.capture_times)}")
            LOGGER.info(f"Weekday capture times: {', '.join(self.capture_times_weekday)}")
            LOGGER.info(f"Weekend capture times: {', '.join(self.capture_times_weekend)}")
        else:
            LOGGER.info("Image capture disabled")

        # Start background tasks
        self._process_task = asyncio.create_task(self._run_process())
        self._send_task = asyncio.create_task(self._run_send())
        if self.include_images:
            self._capture_task = asyncio.create_task(self._run_capture())
    
    async def _run_process(self):
        """Run the workbook processing loop."""
        LOGGER.info(f"Starting process loop for {self.name} (PID: {os.getpid()})")
        try:
            while True:
                current_time = datetime.datetime.now()
                next_process = self._get_next_process_time(current_time)
                sleep_seconds = (next_process - current_time).total_seconds()

                if sleep_seconds <= 0:
                    LOGGER.info(f"Already past process time {next_process.strftime('%H:%M')}, processing now")
                    await self.process_workbook()
                    await asyncio.sleep(1)  # Small gap between checks
                    continue

                LOGGER.info(f"Next process scheduled for {next_process.strftime('%H:%M')} (sleeping {sleep_seconds:.1f} seconds)")
                await asyncio.sleep(sleep_seconds)
                await self.process_workbook()

        except asyncio.CancelledError:
            LOGGER.info(f"Process loop cancelled for {self.name}")
            raise
        except Exception as e:
            LOGGER.error(f"Error in process loop: {e}")
            await asyncio.sleep(60)  # Wait before restarting

    async def _run_send(self):
        """Run the report sending loop."""
        LOGGER.info(f"Starting send loop for {self.name} (PID: {os.getpid()})")
        try:
            while True:
                current_time = datetime.datetime.now()
                next_send = self._get_next_send_time(current_time)
                sleep_seconds = (next_send - current_time).total_seconds()

                if sleep_seconds <= 0:
                    LOGGER.info(f"Already past send time {next_send.strftime('%H:%M')}, sending now")
                    await self.send_report_if_ready()
                    await asyncio.sleep(1)
                    continue

                LOGGER.info(f"Next send scheduled for {next_send.strftime('%H:%M')} (sleeping {sleep_seconds:.1f} seconds)")
                await asyncio.sleep(sleep_seconds)
                await self.send_report_if_ready()

        except asyncio.CancelledError:
            LOGGER.info(f"Send loop cancelled for {self.name}")
            raise
        except Exception as e:
            LOGGER.error(f"Error in send loop: {e}")
            await asyncio.sleep(60)
    
    async def _run_capture(self):
        """Run the image capture loop."""
        LOGGER.info(f"Starting capture loop for {self.name} (PID: {os.getpid()})")
        try:
            while True:
                current_time = datetime.datetime.now()
                next_capture = self._get_next_capture_time(current_time)
                sleep_seconds = (next_capture - current_time).total_seconds()

                if sleep_seconds <= 0:
                    LOGGER.info(f"Already past capture time {next_capture.strftime('%H:%M')}, capturing now")
                    await self.capture_image()
                    await asyncio.sleep(1)
                    continue

                LOGGER.info(f"Next capture scheduled for {next_capture.strftime('%H:%M')} (sleeping {sleep_seconds:.1f} seconds)")
                await asyncio.sleep(sleep_seconds)
                await self.capture_image()

        except asyncio.CancelledError:
            LOGGER.info(f"Capture loop cancelled for {self.name}")
            raise
        except Exception as e:
            LOGGER.error(f"Error in capture loop: {e}")
            await asyncio.sleep(60)

    def _get_next_process_time(self, current_time: datetime.datetime) -> datetime.datetime:
        """Calculate the next process time."""
        today = current_time.date()
        process_dt = datetime.datetime.combine(today, datetime.time(*map(int, self.process_time.split(":"))))
        if current_time > process_dt:
            process_dt += datetime.timedelta(days=1)
        return process_dt

    def _get_next_send_time(self, current_time: datetime.datetime) -> datetime.datetime:
        """Calculate the next send time."""
        today = current_time.date()
        send_dt = datetime.datetime.combine(today, datetime.time(*map(int, self.send_time.split(":"))))
        if current_time > send_dt:
            send_dt += datetime.timedelta(days=1)
        return send_dt

    def _get_next_capture_time(self, current_time: datetime.datetime) -> datetime.datetime:
        """Calculate the next capture time based on weekday/weekend schedules."""
        today = current_time.date()
        tomorrow = today + datetime.timedelta(days=1)
        
        # Determine if today and tomorrow are weekdays
        is_today_weekday = self._is_weekday(today)
        is_tomorrow_weekday = self._is_weekday(tomorrow)
        
        # Get the appropriate capture times for today and tomorrow
        today_capture_times = self.capture_times_weekday if is_today_weekday else self.capture_times_weekend
        tomorrow_capture_times = self.capture_times_weekday if is_tomorrow_weekday else self.capture_times_weekend
        
        # Create datetime objects for today's and tomorrow's capture times
        capture_times_today = [
            datetime.datetime.combine(today, datetime.time(*map(int, t.split(":"))))
            for t in today_capture_times
        ]
        capture_times_tomorrow = [
            datetime.datetime.combine(tomorrow, datetime.time(*map(int, t.split(":"))))
            for t in tomorrow_capture_times
        ]
        
        # Find all future capture times from today and tomorrow
        future_captures = [dt for dt in capture_times_today + capture_times_tomorrow if dt > current_time]
        
        # If we have future captures, return the earliest one
        if future_captures:
            return min(future_captures)
        
        # If no captures today or tomorrow, find the next day
        day_after_tomorrow = tomorrow + datetime.timedelta(days=1)
        is_day_after_tomorrow_weekday = self._is_weekday(day_after_tomorrow)
        next_day_times = self.capture_times_weekday if is_day_after_tomorrow_weekday else self.capture_times_weekend
        
        if next_day_times:
            return datetime.datetime.combine(
                day_after_tomorrow, datetime.time(*map(int, next_day_times[0].split(":")))
            )
        
        # Fallback (shouldn't happen with our defaults)
        return datetime.datetime.combine(
            # Noon the day after tomorrow
            day_after_tomorrow, datetime.time(12, 0)  
        )

    async def capture_image(self):
        """Capture an image from the camera and save it to disk."""
        if not self.include_images or not self.camera_name:
            return

        camera = None
        for name, resource in self.dependencies.items():
            if isinstance(resource, Camera) and self.camera_name.lower() in str(name).lower():
                camera = resource
                LOGGER.info(f"Found camera: {name}")
                break

        if not camera:
            LOGGER.warning(f"Camera '{self.camera_name}' not found in dependencies")
            return

        try:
            LOGGER.info(f"Capturing image from camera '{self.camera_name}'")
            image = await camera.get_image(mime_type="image/jpeg")
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y%m%d_%H%M%S")
            filename = f"{timestamp}_{self.name}.jpg"
            daily_dir = os.path.join(self.images_dir, now.strftime("%Y%m%d"))
            os.makedirs(daily_dir, exist_ok=True)
            image_path = os.path.join(daily_dir, filename)

            if isinstance(image, ViamImage):
                pil_image = Image.open(BytesIO(image.data))
                pil_image.save(image_path, "JPEG")
            elif isinstance(image, bytes):
                with open(image_path, "wb") as f:
                    f.write(image)
            else:
                LOGGER.warning(f"Unsupported image type: {type(image)}")
                return

            self.last_capture_time = now
            LOGGER.info(f"Saved image to {image_path}")
            annotated_path = self.annotate_image(image_path)
            if annotated_path:
                LOGGER.info(f"Annotated image saved to {annotated_path}")

        except Exception as e:
            LOGGER.error(f"Error capturing image: {e}")

    def annotate_image(self, image_path, font_size=20):
        """Annotate an image with timestamp and location information."""
        try:
            img = Image.open(image_path)
            draw = ImageDraw.Draw(img)
            filename = os.path.basename(image_path)
            parts = filename.split('_')
            if len(parts) >= 2 and len(parts[0]) == 8 and len(parts[1]) >= 6:
                date_part, time_part = parts[0], parts[1]
                timestamp_text = f"{date_part[0:4]}-{date_part[4:6]}-{date_part[6:8]} {time_part[0:2]}:{time_part[2:4]}:{time_part[4:6]}"
            else:
                timestamp_text = filename

            text = f"{timestamp_text} - {self.location}"
            font = ImageFont.load_default() if hasattr(ImageFont, 'load_default') else None
            text_bbox = draw.textbbox((0, 0), text, font=font) if font else (0, 0, len(text) * font_size * 0.6, font_size * 1.2)
            text_width, text_height = text_bbox[2] - text_bbox[0], text_bbox[3] - text_bbox[1]
            x, y = img.width - text_width - 10, img.height - text_height - 10

            draw.rectangle([x-5, y-5, x+text_width+5, y+text_height+5], fill=(0, 0, 0, 128))
            draw.text((x, y), text, fill="white", font=font)
            annotated_path = image_path.replace(".jpg", "_annotated.jpg")
            img.save(annotated_path, "JPEG")
            LOGGER.info(f"Created annotated image: {annotated_path}")
            return annotated_path

        except Exception as e:
            LOGGER.error(f"Error annotating image: {e}")
            return image_path

    async def process_workbook(self):
        """Process the Excel workbook for today's data."""
        now = datetime.datetime.now()
        date_str = now.strftime("%Y%m%d")
        try:
            target_date = now.replace(tzinfo=tz.gettz(self.timezone))
            LOGGER.info(f"Processing workbook for date: {date_str}")

            template_path = os.path.join(self.workbooks_dir, "template.xlsx")
            raw_data_path = os.path.join(self.workbooks_dir, "raw_export.xlsx")
            if not os.path.exists(template_path):
                LOGGER.error(f"Template file not found: {template_path}")
                self.workbook_status = "error: missing template"
                self._save_state()
                return

            opening_time, closing_time = self._get_store_hours_for_date(target_date)
            open_hour, open_minute = map(int, opening_time.split(':'))
            close_hour, close_minute = map(int, closing_time.split(':'))
            start_time = target_date.replace(hour=open_hour, minute=open_minute, second=0, microsecond=0)
            end_time = target_date.replace(hour=close_hour, minute=close_minute, second=0, microsecond=0)

            LOGGER.info(f"Exporting data from {start_time} to {end_time}")
            exporter = DataExporter(self.api_key_id, self.api_key, self.org_id, self.location, self.timezone)
            await exporter.export_to_excel(
                raw_data_path, "langer_fill", start_time, end_time,
                bucket_period="PT5M", bucket_method="pct99", include_keys_regex=".*_raw", tab_name="RAW"
            )

            wip_filename = f"{date_str}_{self.name}_wip.xlsx"
            final_filename = f"{date_str}_{self.name}.xlsx"
            wip_path = os.path.join(self.workbooks_dir, wip_filename)
            final_path = os.path.join(self.workbooks_dir, final_filename)

            shutil.copy(template_path, wip_path)
            num_data_rows = self._update_raw_import_sheet(raw_data_path, wip_path)
            LOGGER.info(f"Updated Raw Import sheet with {num_data_rows} rows")
            self._fix_workbook(wip_path, num_data_rows, final_path)
            LOGGER.info(f"Created final workbook: {final_path}")

            if os.path.exists(wip_path):
                os.remove(wip_path)
                LOGGER.info(f"Removed temporary WIP file: {wip_path}")

            self.last_workbook_path = final_path
            self.last_processed_time = now
            self.workbook_status = "processed"
            self._save_state()

        except Exception as e:
            LOGGER.error(f"Failed to process workbook: {e}")
            self.workbook_status = f"error: {str(e)}"
            self._save_state()

    def _get_store_hours_for_date(self, date):
        """Get store hours for the specified date."""
        return tuple(self.hours_weekends if date.weekday() >= 5 else self.hours_weekdays)
    
    def _is_weekday(self, date: datetime.date) -> bool:
        """Check if the given date is a weekday (0=Monday, 6=Sunday)."""
        # 0-4 are the weekdays (Monday-Friday)
        return date.weekday() < 5

    def _update_raw_import_sheet(self, raw_file, output_file):
        """Update the Raw Import sheet in the output workbook."""
        try:
            LOGGER.info(f"Loading raw data from {raw_file}")
            raw_wb = openpyxl.load_workbook(raw_file, data_only=True)
            if "RAW" not in raw_wb.sheetnames:
                raise ValueError("RAW sheet not found in exported data")
            raw_sheet = raw_wb["RAW"]
            data_rows = list(raw_sheet.iter_rows(min_row=2, values_only=True))
            LOGGER.info(f"Loaded {len(data_rows)} rows from raw export")

            LOGGER.info(f"Opening output workbook: {output_file}")
            output_wb = openpyxl.load_workbook(output_file)
            if "Raw Import" not in output_wb.sheetnames:
                raise ValueError("Raw Import sheet not found in template")
            output_sheet = output_wb["Raw Import"]

            LOGGER.info("Clearing existing data from Raw Import sheet")
            for row in output_sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None

            LOGGER.info("Copying data to Raw Import sheet")
            for r_idx, row_data in enumerate(data_rows, start=2):
                for c_idx, value in enumerate(row_data, start=1):
                    output_sheet.cell(row=r_idx, column=c_idx).value = value

            LOGGER.info(f"Saving updated workbook to {output_file}")
            output_wb.save(output_file)
            LOGGER.info(f"Raw Import sheet updated with {len(data_rows)} rows")
            return len(data_rows)

        except Exception as e:
            LOGGER.error(f"Error updating Raw Import sheet: {e}")
            raise

    def _get_sheet_mappings(self, excel_path):
        """Extract sheet mappings from the Excel workbook."""
        temp_dir = os.path.join(self.workbooks_dir, "temp_excel")
        os.makedirs(temp_dir, exist_ok=True)
        try:
            with zipfile.ZipFile(excel_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)

            workbook_xml_path = os.path.join(temp_dir, "xl", "workbook.xml")
            rels_xml_path = os.path.join(temp_dir, "xl", "_rels", "workbook.xml.rels")
            if not os.path.exists(workbook_xml_path) or not os.path.exists(rels_xml_path):
                raise FileNotFoundError(f"Required XML files not found in {excel_path}")

            sheet_mapping = {}
            wb_tree = ET.parse(workbook_xml_path)
            wb_root = wb_tree.getroot()
            ns = {
                'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            # Use fully qualified namespace for r:id
            sheet_rel_map = {}
            for sheet in wb_root.findall(".//ns:sheets/ns:sheet", ns):
                sheet_name = sheet.attrib.get("name", "unknown")
                sheet_rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                if sheet_rel_id:
                    sheet_rel_map[sheet_rel_id] = sheet_name
                else:
                    LOGGER.warning(f"Sheet '{sheet_name}' missing r:id attribute in {workbook_xml_path}")

            rels_tree = ET.parse(rels_xml_path)
            rels_root = rels_tree.getroot()
            rels_ns = {'ns': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            for rel in rels_root.findall(".//ns:Relationship", rels_ns):
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rel_id in sheet_rel_map and "worksheets" in target:
                    sheet_mapping[sheet_rel_map[rel_id]] = os.path.basename(target)

            LOGGER.info(f"Sheet mappings: {sheet_mapping}")
            return sheet_mapping

        except Exception as e:
            LOGGER.error(f"Error extracting sheet mappings: {e}")
            raise
        finally:
            if os.path.exists(temp_dir) and "temp_excel" not in excel_path:
                shutil.rmtree(temp_dir)

    def _fix_workbook(self, wip_path, num_data_rows, final_path):
        """Fix the workbook structure for row counts and formulas."""
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        temp_dir = os.path.join(self.workbooks_dir, f"temp_excel_{timestamp}")
        try:
            if not os.path.exists(wip_path):
                raise FileNotFoundError(f"WIP file not found: {wip_path}")

            os.makedirs(temp_dir, exist_ok=True)
            LOGGER.info(f"Created temp directory: {temp_dir}")
            with zipfile.ZipFile(wip_path, "r") as zip_ref:
                zip_ref.extractall(temp_dir)

            sheet_mappings = self._get_sheet_mappings(wip_path)
            namespaces = {
                'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
            }
            for prefix, uri in namespaces.items():
                ET.register_namespace(prefix, uri)
            ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')

            worksheets_dir = os.path.join(temp_dir, "xl", "worksheets")
            if not os.path.exists(worksheets_dir):
                raise FileNotFoundError(f"Worksheets directory not found: {worksheets_dir}")

            sheets_to_process = ["Calibrated Values", "Bounded Calibrated", "Empty Shelf Tracker"]
            for sheet_name in sheets_to_process:
                if sheet_name not in sheet_mappings:
                    LOGGER.warning(f"Sheet '{sheet_name}' not found in workbook. Skipping...")
                    continue
                sheet_xml_path = os.path.join(worksheets_dir, sheet_mappings[sheet_name])
                if not os.path.exists(sheet_xml_path):
                    LOGGER.error(f"Sheet XML file not found: {sheet_xml_path}")
                    continue

                LOGGER.info(f"Processing sheet: {sheet_name}")
                tree = ET.parse(sheet_xml_path)
                root = tree.getroot()
                sheet_data = root.find(".//ns:sheetData", namespaces)
                if sheet_data is None:
                    LOGGER.warning(f"No sheetData found in {sheet_name}, skipping modifications")
                    continue

                rows_to_remove = [row for row in sheet_data.findall(".//ns:row", namespaces) if int(row.attrib.get("r", "0")) > num_data_rows + 1]
                if rows_to_remove:
                    first_row = rows_to_remove[0].attrib.get('r', "N/A")
                    last_row = rows_to_remove[-1].attrib.get('r', "N/A")
                    for row in rows_to_remove:
                        sheet_data.remove(row)
                    LOGGER.info(f"Removed {len(rows_to_remove)} excess rows ({first_row} to {last_row}) from {sheet_name}")

                tree.write(sheet_xml_path, encoding="UTF-8", xml_declaration=True)
                LOGGER.info(f"Saved modifications to {sheet_xml_path}")

            LOGGER.info(f"Creating final workbook: {final_path}")
            with zipfile.ZipFile(final_path, "w", zipfile.ZIP_DEFLATED) as zip_out:
                for root_dir, _, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        zip_out.write(file_path, arcname)

            LOGGER.info(f"Successfully created final workbook: {final_path}")
            return final_path

        except Exception as e:
            LOGGER.error(f"Error fixing workbook: {e}")
            raise
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                LOGGER.info(f"Cleaned up temporary directory: {temp_dir}")

    def _get_daily_images(self, day_str):
        """Get all captured images for a specific day."""
        daily_dir = os.path.join(self.images_dir, day_str)
        if not os.path.exists(daily_dir):
            LOGGER.info(f"No image directory for {day_str}")
            return []
        image_files = sorted(
            [os.path.join(daily_dir, f) for f in os.listdir(daily_dir) if f.endswith(".jpg") and not f.endswith("_annotated.jpg")]
        )
        LOGGER.info(f"Found {len(image_files)} images for {day_str}")
        return image_files

    async def send_report_if_ready(self):
        """Send a report if a processed workbook is available."""
        if not self.last_workbook_path or not os.path.exists(self.last_workbook_path):
            LOGGER.error("No processed workbook available to send report")
            self.report_status = "error: no processed workbook"
            self._save_state()
            return

        now = datetime.datetime.now()
        date_str = now.strftime("%Y%m%d")
        try:
            daily_images = []
            if self.include_images:
                image_files = self._get_daily_images(date_str)
                for img_path in image_files:
                    try:
                        annotated_path = self.annotate_image(img_path)
                        if annotated_path:
                            daily_images.append(annotated_path)
                    except Exception as e:
                        LOGGER.error(f"Error annotating image {img_path}: {e}")

            await self.send_report(self.last_workbook_path, daily_images)
            self.last_sent_time = now
            self.total_reports_sent += 1
            self.report_status = "sent"
            self._save_state()
            LOGGER.info(f"Sent email report for {date_str} with {len(daily_images)} images")

        except Exception as e:
            self.report_status = f"error: {str(e)}"
            LOGGER.error(f"Failed to send report for {date_str}: {e}")
            self._save_state()

    async def send_report(self, workbook_path, image_paths=None):
        """Send the report via email with images and workbook."""
        if not self.sendgrid_api_key:
            LOGGER.error("No SendGrid API key configured")
            return

        try:
            LOGGER.info(f"Preparing report email with workbook: {os.path.basename(workbook_path)}")
            now = datetime.datetime.now()
            subject = f"Daily Report: {now.strftime('%Y-%m-%d')} - {self.location}"
            body_text = f"The Excel workbook is attached with data for review.\nLocation: {self.location}\n"
            if image_paths:
                body_text += f"Also attached are {len(image_paths)} images captured during the day.\n"
            if self.teleop_url and self.teleop_url != "#":
                body_text += f"Click here for the link to a real-time view of the store: {self.teleop_url}"

            html_content = f"""<html><body><p>The Excel workbook is attached with data for review.</p><p>Location: {self.location}</p>"""
            if image_paths:
                html_content += f"<p>Also attached are {len(image_paths)} images captured during the day.</p>"
            if self.teleop_url and self.teleop_url != "#":
                html_content += f"""<p>Click <a href="{self.teleop_url}">here</a> for the link to a real-time view of the store.</p>"""
            html_content += "</body></html>"

            valid_recipients = [r for r in self.recipients if isinstance(r, str) and '@' in r]
            if not valid_recipients:
                LOGGER.error("No valid recipients found")
                return

            message = Mail(
                from_email=Email(self.sender_email, self.sender_name),
                to_emails=valid_recipients,
                subject=subject,
                plain_text_content=Content("text/plain", body_text)
            )
            message.add_content(Content("text/html", html_content))

            if image_paths:
                def sort_by_timestamp(path):
                    parts = os.path.basename(path).split('_')
                    if len(parts) >= 2:
                        try:
                            return datetime.datetime.strptime(parts[0] + '_' + parts[1], "%Y%m%d_%H%M%S")
                        except ValueError:
                            return path
                    return path

                sorted_images = sorted(image_paths, key=sort_by_timestamp, reverse=True)
                for img_path in sorted_images:
                    try:
                        with open(img_path, "rb") as f:
                            img_content = base64.b64encode(f.read()).decode()
                        img_name = os.path.basename(img_path)
                        img_attachment = Attachment(
                            FileContent(img_content), FileName(img_name), FileType("image/jpeg"), Disposition("attachment")
                        )
                        message.add_attachment(img_attachment)
                        LOGGER.info(f"Added image attachment to report: {img_name}")
                    except Exception as e:
                        LOGGER.error(f"Error attaching image: {e}")

            with open(workbook_path, "rb") as f:
                wb_content = base64.b64encode(f.read()).decode()
            wb_name = os.path.basename(workbook_path)
            wb_attachment = Attachment(
                FileContent(wb_content), FileName(wb_name),
                FileType("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"), Disposition("attachment")
            )
            message.add_attachment(wb_attachment)
            LOGGER.info(f"Added workbook attachment to report: {wb_name}")

            LOGGER.info(f"Sending email report to {len(valid_recipients)} recipients")
            sg = SendGridAPIClient(self.sendgrid_api_key)
            response = sg.send(message)
            LOGGER.info(f"Email sent via SendGrid API. Status code: {response.status_code}")

        except Exception as e:
            LOGGER.error(f"Failed to send report: {e}")

    async def get_readings(self, *, extra: Optional[Dict[str, Any]] = None, **kwargs) -> Dict[str, SensorReading]:
        """Get current sensor readings."""
        now = datetime.datetime.now()
        is_today_weekday = self._is_weekday(now.date())
        
        next_process = self._get_next_process_time(now)
        next_send = self._get_next_send_time(now)
        next_capture = self._get_next_capture_time(now) if self.include_images else None

        readings = {
            "location": self.location,
            "last_processed_time": str(self.last_processed_time) if self.last_processed_time else "never",
            "last_sent_time": str(self.last_sent_time) if self.last_sent_time else "never",
            "last_workbook_path": self.last_workbook_path or "none",
            "total_reports_sent": self.total_reports_sent,
            "report_status": self.report_status,
            "workbook_status": self.workbook_status,
            "next_process_time": str(next_process),
            "next_send_time": str(next_send),
            "timezone": self.timezone,
            "hours_weekdays": self.hours_weekdays,
            "hours_weekends": self.hours_weekends,
            "current_day_type": "weekday" if is_today_weekday else "weekend",
            "pid": os.getpid()
        }

        if self.include_images:
            readings.update({
                "include_images": True,
                "camera_name": self.camera_name,
                "capture_times_weekday": self.capture_times_weekday,
                "capture_times_weekend": self.capture_times_weekend,
                "current_capture_times": self.capture_times_weekday if is_today_weekday else self.capture_times_weekend,
                "last_capture_time": str(self.last_capture_time) if self.last_capture_time else "never",
                "next_capture_time": str(next_capture) if next_capture else "none scheduled"
            })
        else:
            readings["include_images"] = False

        return readings

    async def do_command(self, command: Dict[str, Any], *, timeout: Optional[float] = None, **kwargs) -> Dict[str, Any]:
        """Handle custom commands."""
        cmd = command.get("command", "")

        if cmd == "process_and_send":
            date = command.get("date", datetime.datetime.now().strftime("%Y%m%d"))
            try:
                timestamp = datetime.datetime.strptime(day, "%Y%m%d")
                await self.process_workbook()
                await self.send_report_if_ready()
                return {"status": "completed", "message": f"Processed and sent report for {day}"}
            except ValueError:
                return {"status": "error", "message": f"Invalid day format: {day}, use YYYYMMDD"}

        elif cmd == "process":
            date = command.get("date", datetime.datetime.now().strftime("%Y%m%d"))
            try:
                await self.process_workbook()
                return {"status": "completed", "message": f"Processed workbook for {day}", "path": self.last_workbook_path}
            except ValueError:
                return {"status": "error", "message": f"Invalid day format: {day}, use YYYYMMDD"}

        elif cmd == "capture_image":
            if not self.include_images:
                return {"status": "error", "message": "Image capture not enabled"}
            await self.capture_image()
            return {"status": "completed", "message": f"Captured image at {datetime.datetime.now()}"}

        elif cmd == "test_email":
            if not self.sendgrid_api_key:
                return {"status": "error", "message": "No SendGrid API key configured"}
            try:
                timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                subject = f"Test Report from {self.location}"
                body = f"This is a test report from {self.name} at {self.location}.\nTime: {timestamp}"
                message = Mail(
                    from_email=Email(self.sender_email, self.sender_name),
                    to_emails=self.recipients,
                    subject=subject,
                    plain_text_content=Content("text/plain", body)
                )
                html_body = body.replace("\n", "<br>")
                message.add_content(Content("text/html", f"<html><body><p>{html_body}</p></body></html>"))
                sg = SendGridAPIClient(self.sendgrid_api_key)
                response = sg.send(message)
                return {"status": "completed", "message": f"Test report sent with status code {response.status_code}"}
            except Exception as e:
                return {"status": "error", "message": f"Failed to send test report: {str(e)}"}

        elif cmd == "get_schedule":
            now = datetime.datetime.now()
            is_today_weekday = self._is_weekday(now.date())
            return {
                "status": "completed",
                "process_time": self.process_time,
                "send_time": self.send_time,
                "next_process": str(self._get_next_process_time(now)),
                "next_send": str(self._get_next_send_time(now)),
                "current_day_type": "weekday" if is_today_weekday else "weekend",
                "capture_times_weekday": self.capture_times_weekday if self.include_images else [],
                "capture_times_weekend": self.capture_times_weekend if self.include_images else [],
                "current_capture_times": (self.capture_times_weekday if is_today_weekday else self.capture_times_weekend) if self.include_images else [],
                "next_capture": str(self._get_next_capture_time(now)) if self.include_images else "none scheduled"
            }

        return {"status": "error", "message": f"Unknown command: {cmd}"}