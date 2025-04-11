import asyncio
import logging
import re
import pytz
import numpy as np
from datetime import datetime, timedelta
from viam.app.viam_client import ViamClient, DataClient
from viam.rpc.dial import DialOptions, Credentials
from viam.proto.app.data import Filter, Order
from google.protobuf.timestamp_pb2 import Timestamp
from openpyxl import Workbook

LOGGER = logging.getLogger(__name__)

class DataExporter:
    """
    A simplified version of vde.py data exporter, specifically tailored for
    exporting workbook data from the Viam Data API, now with image retrieval support.
    """
    
    def __init__(self, api_key_id, api_key, org_id, location_id, timezone="America/New_York"):
        """
        Initialize the data exporter.
        
        Args:
            api_key_id: Viam API key ID
            api_key: Viam API key
            org_id: Viam organization ID
            location_id: Viam location ID
            timezone: Timezone for timestamps (default: America/New_York)
        """
        self.api_key_id = api_key_id
        self.api_key = api_key
        self.org_id = org_id
        self.location_id = location_id  # Add location_id
        self.timezone = self._parse_timezone(timezone)
        self.data_client = None  # Will be set when connected
        
    def _parse_timezone(self, tz_str):
        """Convert a timezone string to a pytz timezone object."""
        return pytz.timezone(tz_str)
        
    async def connect(self):
        """Connect to the Viam API."""
        LOGGER.info("Connecting to Viam API")
        dial_options = DialOptions(
            credentials=Credentials(
                type="api-key",
                payload=self.api_key,
            ),
            auth_entity=self.api_key_id
        )
        client = await ViamClient.create_from_dial_options(dial_options)
        self.data_client = client.data_client  # Store data_client for reuse
        return client
    
    async def export_to_excel(self, 
                            output_file, 
                            resource_name, 
                            start_time, 
                            end_time, 
                            bucket_period="PT5M", 
                            bucket_method="pct99", 
                            include_keys_regex=".*_raw",
                            tab_name="RAW"):
        """
        Export data to an Excel file.
        
        Args:
            output_file: Path to save the Excel file
            resource_name: The Viam resource name to query
            start_time: Start time for the data query
            end_time: End time for the data query
            bucket_period: Time bucket period (ISO8601 duration string or timedelta)
            bucket_method: Aggregation method for bucketing
            include_keys_regex: Regex pattern for keys to include
            tab_name: Name of the worksheet tab
            
        Returns:
            Path to the created Excel file
        """
        LOGGER.info(f"Exporting data from {start_time} to {end_time}")
        
        # Parse bucket_period if it's a string
        if isinstance(bucket_period, str):
            try:
                from isodate import parse_duration
                bucket_period = parse_duration(bucket_period)
            except ImportError:
                LOGGER.warning("isodate package not available, using default bucket period")
                bucket_period = timedelta(minutes=5)
            
        # Connect to Viam API
        client = await self.connect()
        try:
            # Build the filter pipeline
            match_predicate = {
                "organization_id": self.org_id,
                "component_name": resource_name,
                "time_received": {
                    "$gte": start_time,
                    "$lt": end_time
                }
            }
            
            pipeline = [
                {"$match": match_predicate},
                {"$sort": {"time_received": 1}}
            ]
            
            # Fetch all data with pagination
            all_data = []
            skip = 0
            limit = 1000
            
            while True:
                LOGGER.info(f"Retrieving data from {skip} to {skip + limit}")
                
                # Clone the pipeline and add pagination
                batch_pipeline = pipeline.copy()
                batch_pipeline.append({"$skip": skip})
                batch_pipeline.append({"$limit": limit})
                
                LOGGER.debug(f"Executing pipeline: {batch_pipeline}")
                batch = await self.data_client.tabular_data_by_mql(organization_id=self.org_id, query=batch_pipeline)
                
                batch_len = len(batch)
                if batch_len == 0:
                    break
                    
                all_data.extend(batch)
                LOGGER.info(f"Retrieved {batch_len} records")
                
                if batch_len < limit:  # Less than limit means we've reached the end
                    break
                    
                skip += limit
            
            # Process data (bucket if needed)
            if bucket_period:
                all_data = self._bucket_data(all_data, bucket_period, bucket_method, include_keys_regex)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = tab_name
            
            # Write headers
            if all_data:
                data_keys = sorted(all_data[0]["data"]["readings"].keys())
                headers = ["time_received"] + data_keys
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # Write data rows
                for row_idx, row in enumerate(all_data, 2):
                    time_received = row["time_received"]
                    utc_time = time_received.replace(tzinfo=pytz.utc)
                    local_time = utc_time.astimezone(tz=self.timezone)
                    
                    # Write timestamp
                    ws.cell(row=row_idx, column=1, value=local_time.replace(tzinfo=None))
                    
                    # Write data values
                    for col_idx, key in enumerate(data_keys, 2):
                        try:
                            if key in row["data"]["readings"]:
                                value = row["data"]["readings"][key]
                                ws.cell(row=row_idx, column=col_idx, value=value)
                        except Exception as e:
                            LOGGER.warning(f"Error writing value for key {key}: {e}")
            
            # Save workbook
            wb.save(output_file)
            LOGGER.info(f"Saved data to {output_file} with {len(all_data)} rows")
            
            return output_file
            
        finally:
            client.close()
    
    async def get_closest_images(self, component_name, start_time, end_time, desired_times):
        """
        Retrieve images closest to the specified times from Viam Data Management.
        
        Args:
            component_name: Name of the component (e.g., "ffmpeg")
            start_time: Start of the time range (datetime with timezone)
            end_time: End of the time range (datetime with timezone)
            desired_times: List of datetime objects with timezone for desired image times
            
        Returns:
            List of tuples: [(desired_time, BinaryData or None), ...]
        """
        from viam.proto.app.data import CaptureInterval, Order, BinaryID

        # Initialize default result with None values
        result = [(dt, None) for dt in desired_times]
        
        if self.data_client is None:
            LOGGER.error("Data client is None, cannot retrieve images")
            return result
            
        try:
            # Convert times to UTC for API consistency
            start_time_utc = start_time.astimezone(pytz.utc)
            end_time_utc = end_time.astimezone(pytz.utc)
            desired_times_utc = [dt.astimezone(pytz.utc) for dt in desired_times]

            # Create filter with CaptureInterval for time range
            start_ts = Timestamp()
            start_ts.FromDatetime(start_time_utc)
            end_ts = Timestamp()
            end_ts.FromDatetime(end_time_utc)
            capture_interval = CaptureInterval(start=start_ts, end=end_ts)
            filter = Filter(
                component_name=component_name,
                method="ReadImage",
                interval=capture_interval,
                organization_ids=[self.org_id]
            )

            # Query image metadata, sorted by time (ascending)
            LOGGER.info(f"Querying binary data with filter: component_name={component_name}, method=ReadImage, interval={start_time_utc} to {end_time_utc}, organization_ids={self.org_id}")
            
            try:
                all_images, _, _ = await self.data_client.binary_data_by_filter(
                    filter=filter,
                    include_binary_data=False,
                    sort_order=Order.ORDER_ASCENDING
                )
            except Exception as e:
                LOGGER.error(f"Failed to fetch image metadata: {e}")
                return result

            if not all_images:
                LOGGER.warning(f"No images found for {component_name} between {start_time} and {end_time}")
                return result

            # Find closest image for each desired time
            binary_ids = []
            valid_indices = []
            
            for i, dt_utc in enumerate(desired_times_utc):
                try:
                    closest_image = min(all_images, key=lambda img: abs(img.metadata.time_received.ToDatetime().replace(tzinfo=pytz.utc) - dt_utc))
                    
                    # Create a BinaryID object with all required fields
                    try:
                        binary_id = BinaryID(
                            file_id=closest_image.metadata.id,
                            organization_id=self.org_id,
                            location_id=self.location_id
                        )
                        binary_ids.append(binary_id)
                        valid_indices.append(i)
                    except Exception as e:
                        LOGGER.error(f"Error creating BinaryID for image at time {dt_utc}: {e}")
                except Exception as e:
                    LOGGER.error(f"Failed to find closest image for {dt_utc}: {e}")

            if not binary_ids:
                LOGGER.warning("No valid binary IDs found")
                return result
                
            # Retrieve binary data for selected images
            LOGGER.info(f"Retrieving binary data for {len(binary_ids)} image IDs")
            
            try:
                binary_data_list = await self.data_client.binary_data_by_ids(binary_ids)
                
                # Map binary data back to desired times
                for i, idx in enumerate(valid_indices):
                    if i < len(binary_data_list):
                        result[idx] = (desired_times[idx], binary_data_list[i])
                        
                LOGGER.info(f"Retrieved {len(binary_data_list)} images")
                return result
                
            except Exception as e:
                LOGGER.error(f"Failed to retrieve binary data: {e}")
                return result
                
        except Exception as e:
            LOGGER.error(f"Error in get_closest_images: {e}")
            return result

    def _floor_timestamp(self, ts, bucket_td):
        """Round a timestamp down to the nearest bucket interval."""
        epoch = datetime(1970, 1, 1, tzinfo=ts.tzinfo)
        bucket_count = (ts - epoch) // bucket_td
        return epoch + bucket_count * bucket_td
    
    def _bucket_data(self, data, bucket_period, bucket_method, include_keys_regex=None):
        """
        Bucket data by the specified time period and aggregation method.
        
        Args:
            data: List of data points
            bucket_period: Timedelta object specifying the bucket size
            bucket_method: Aggregation method (min, max, avg, first, last, pct95, pct99)
            include_keys_regex: Regex pattern for keys to include
        
        Returns:
            List of aggregated data points
        """
        LOGGER.info(f"Bucketing data with period {bucket_period} using method {bucket_method}")
        
        include_regex = None
        if include_keys_regex:
            include_regex = re.compile(include_keys_regex)
        
        # Group data by time bucket
        bucketed_data = {}
        for row in data:
            time_received = row["time_received"]
            bucket = self._floor_timestamp(time_received, bucket_period)
            
            if bucket not in bucketed_data:
                bucketed_data[bucket] = {}
                
            for key, value in row["data"]["readings"].items():
                # Apply key filtering if regex provided
                if include_regex and not include_regex.match(key):
                    continue
                    
                if key not in bucketed_data[bucket]:
                    bucketed_data[bucket][key] = []
                    
                bucketed_data[bucket][key].append(value)
        
        LOGGER.debug(f"Created {len(bucketed_data)} time buckets")
        
        # Aggregate data in each bucket
        aggregated_data = []
        for bucket, readings in bucketed_data.items():
            aggregated_reading = {}
            
            for key, values in readings.items():
                if bucket_method == "max":
                    aggregated_reading[key] = max(values)
                elif bucket_method == "min":
                    aggregated_reading[key] = min(values)
                elif bucket_method == "avg":
                    aggregated_reading[key] = sum(values) / len(values)
                elif bucket_method == "first":
                    aggregated_reading[key] = values[0]
                elif bucket_method == "last":
                    aggregated_reading[key] = values[-1]
                elif bucket_method == "pct95":
                    aggregated_reading[key] = np.percentile(values, 95)
                elif bucket_method == "pct99":
                    aggregated_reading[key] = np.percentile(values, 99)
                else:
                    LOGGER.warning(f"Unsupported bucket method: {bucket_method}, using max")
                    aggregated_reading[key] = max(values)
            
            aggregated_data.append({
                "time_received": bucket,
                "data": {"readings": aggregated_reading}
            })
        
        # Sort by time
        aggregated_data.sort(key=lambda x: x["time_received"])
        return aggregated_data